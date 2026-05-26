import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter

REPORT_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "reports")

HEADER_COLOR_US = "1F4E79"   # dark blue
HEADER_COLOR_HK = "375623"   # dark green
HEADER_COLOR_SUM = "4A235A"  # dark purple

_thin = Side(style="thin", color="CCCCCC")
_border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)


def _hfill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _color_score(score) -> str:
    """Return hex fill colour based on score magnitude."""
    try:
        v = float(score)
    except (TypeError, ValueError):
        return "FFFFFF"
    if v >= 0.4:
        return "C6EFCE"   # green
    if v >= 0.1:
        return "FFEB9C"   # yellow
    if v >= -0.1:
        return "FFFFFF"   # neutral
    if v >= -0.4:
        return "FFCC99"   # orange
    return "FFC7CE"       # red


def _write_sheet(ws, rows: list[dict], header_hex: str) -> None:
    if not rows:
        return

    cols = list(rows[0].keys())
    header_fill  = _hfill(header_hex)
    header_font  = Font(bold=True, color="FFFFFF", size=10)
    center       = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_wrap    = Alignment(horizontal="left",   vertical="center", wrap_text=True)

    # Write header row
    for c_idx, col in enumerate(cols, 1):
        cell = ws.cell(row=1, column=c_idx, value=col)
        cell.fill   = header_fill
        cell.font   = header_font
        cell.border = _border
        cell.alignment = center

    ws.row_dimensions[1].height = 32

    score_cols = {"Money Flow Score", "Direction Score", "Composite Score"}

    for r_idx, row in enumerate(rows, 2):
        row_fill = PatternFill("solid", fgColor="F2F2F2") if r_idx % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")

        for c_idx, col in enumerate(cols, 1):
            val  = row.get(col, "")
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.border = _border

            if col in score_cols:
                cell.fill      = PatternFill("solid", fgColor=_color_score(val))
                cell.alignment = center
                cell.font      = Font(size=9, bold=True)
            elif col in ("Rationale", "Legs"):
                cell.alignment = left_wrap
                cell.fill      = row_fill
                cell.font      = Font(size=9)
            elif col in ("Strategy", "Bias", "Vol Regime"):
                cell.alignment = center
                cell.fill      = row_fill
                cell.font      = Font(size=9, bold=True)
            else:
                cell.alignment = center
                cell.fill      = row_fill
                cell.font      = Font(size=9)

    # Auto-size columns
    wide_cols = {"Rationale", "Legs"}
    for c_idx, col in enumerate(cols, 1):
        if col in wide_cols:
            ws.column_dimensions[get_column_letter(c_idx)].width = 52
        else:
            max_len = max(
                (len(str(rows[r].get(col, ""))) for r in range(len(rows))),
                default=0,
            )
            ws.column_dimensions[get_column_letter(c_idx)].width = max(len(col) + 4, min(max_len + 2, 40))

    ws.freeze_panes = "A2"


def save_report(results: list[dict], filename: str = "TradingReport.xlsx") -> str:
    os.makedirs(REPORT_DIR, exist_ok=True)
    path = os.path.join(REPORT_DIR, filename)

    us_rows = [r for r in results if r.get("Market") == "US"]
    hk_rows = [r for r in results if r.get("Market") == "HK"]

    wb = Workbook()

    # Summary sheet (all markets)
    ws_sum = wb.active
    ws_sum.title = "Summary"
    _write_sheet(ws_sum, results, HEADER_COLOR_SUM)

    # Per-market sheets
    if us_rows:
        ws_us = wb.create_sheet("US Top 5")
        _write_sheet(ws_us, us_rows, HEADER_COLOR_US)

    if hk_rows:
        ws_hk = wb.create_sheet("HK Top 5")
        _write_sheet(ws_hk, hk_rows, HEADER_COLOR_HK)

    try:
        wb.save(path)
        print(f"\nReport saved: {path}")
    except PermissionError:
        print(f"\nERROR: Cannot save report — please close {path} if it is open in Excel.")
        return path

    return path


def print_summary(results: list[dict]) -> None:
    """Print a compact terminal summary table."""
    cols = ["Market", "Code", "Name", "Price", "MFI", "CMF", "RSI",
            "Composite Score", "Strategy", "Bias"]

    header = f"{'Market':<6} {'Code':<12} {'Name':<20} {'Price':>8} {'MFI':>6} {'CMF':>7} {'RSI':>6} {'Score':>7}  {'Bias':<10} Strategy"
    print("\n" + "-" * 110)
    print(header)
    print("-" * 110)

    for r in results:
        print(
            f"{r.get('Market',''):<6} "
            f"{r.get('Code',''):<12} "
            f"{str(r.get('Name',''))[:19]:<20} "
            f"{r.get('Price', 0):>8.2f} "
            f"{r.get('MFI', 0):>6.1f} "
            f"{r.get('CMF', 0):>7.4f} "
            f"{(r.get('RSI') or 0):>6.1f} "
            f"{r.get('Composite Score', 0):>7.4f}  "
            f"{r.get('Bias', ''):<10} "
            f"{r.get('Strategy', '')}"
        )

    print("-" * 110)
