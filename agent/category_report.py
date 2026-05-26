"""
category_report.py — Category P&L console summary and Excel report.

Console summary: printed after every monitor cycle (lightweight, no API calls).
Excel report:    generated end-of-day or on demand; 4 sheets with color coding.
"""
from __future__ import annotations

import json
import os
from datetime import date, datetime

from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

from agent.category_tracker import (
    CATEGORIES, classify_trade, compute_category_stats, compute_monthly_trend,
)
from agent.position_manager import get_all_trades
from agent.constants import STOCK_TP_PCT, STOCK_CL_PCT, TP_CREDIT_REMAINING, CL_CREDIT_MULT, HKD_USD_RATE

_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
_REPORTS_DIR   = os.path.join(_ROOT, "reports")
_HISTORY_FILE  = os.path.join(_ROOT, "pnl_history.json")

# ── Colour palette ─────────────────────────────────────────────────────────────
_CLR_HEADER      = "2D3748"   # dark slate
_CLR_TOTAL       = "4A5568"   # medium slate
_CLR_ALT_ROW     = "F7FAFC"   # very light blue-grey
_CLR_GREEN_DARK  = "276749"   # text on gain cells
_CLR_GREEN_FILL  = "C6F6D5"   # light green fill
_CLR_RED_DARK    = "9B2C2C"   # text on loss cells
_CLR_RED_FILL    = "FED7D7"   # light red fill
_CLR_WIN         = "C6F6D5"
_CLR_LOSS        = "FED7D7"
_CLR_BREAK       = "FEFCBF"   # pale yellow for $0


def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _border() -> Border:
    thin = Side(style="thin", color="CBD5E0")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def _pnl_fill(value: float) -> PatternFill | None:
    if value > 0:
        return _fill(_CLR_GREEN_FILL)
    if value < 0:
        return _fill(_CLR_RED_FILL)
    return None


def _fmt_money(v: float) -> str:
    sign = "+" if v >= 0 else ""
    return f"{sign}${v:,.2f}"


def _fmt_pct(v: float | None) -> str:
    if v is None:
        return "n/a"
    return f"{v * 100:.1f}%"


# ── Position P&L fetch (standalone mode) ──────────────────────────────────────

def _fetch_pos_pnl(account_us: dict | None,
                   account_hk: dict | None = None) -> dict[str, float]:
    """IBKR positions carry no live P&L — returns empty dict."""
    return {}


# ── PnL history persistence ────────────────────────────────────────────────────

def _append_pnl_history(stats: dict[str, dict], report_date: date) -> None:
    """Upsert today's category stats into pnl_history.json."""
    try:
        if os.path.exists(_HISTORY_FILE):
            with open(_HISTORY_FILE, encoding="utf-8") as f:
                data = json.load(f)
        else:
            data = {"snapshots": []}

        date_s   = report_date.isoformat()
        snapshot = {"date": date_s}
        for cat, s in stats.items():
            snapshot[cat] = {
                "open":             s["open"],
                "closed_alltime":   s["closed_alltime"],
                "win_rate":         round(s["win_rate"], 4) if s["win_rate"] is not None else None,
                "realized_today":   round(s["realized_today"], 2),
                "realized_alltime": round(s["realized_alltime"], 2),
                "unrealized":       round(s["unrealized"], 2),
            }

        # Upsert: replace existing entry for same date, otherwise append
        snapshots = data.get("snapshots", [])
        idx = next((i for i, s in enumerate(snapshots) if s.get("date") == date_s), None)
        if idx is not None:
            snapshots[idx] = snapshot
        else:
            snapshots.append(snapshot)
        data["snapshots"] = snapshots

        tmp = _HISTORY_FILE + ".tmp"
        with open(tmp, "w", encoding="utf-8") as f:
            json.dump(data, f, indent=2, default=str)
        os.replace(tmp, _HISTORY_FILE)
    except Exception as e:
        print(f"  [category_report] WARNING: could not write pnl_history.json: {e}")


# ── Console summary ────────────────────────────────────────────────────────────

def print_category_summary(stats: dict[str, dict]) -> None:
    """Print the category P&L table to stdout (ASCII — safe on all platforms)."""
    W = 88
    print("\n" + "=" * W)
    print(f"  All values in USD  (HK Stocks converted at {HKD_USD_RATE:.2f} HKD/USD)")
    print(f"  {'Category':<18} {'Open':>5}  {'Closed':>6}  {'Win%':>6}  "
          f"{'Real.(Today)':>13}  {'Real.(All-Time)':>15}  {'Unrealized':>11}")
    print("  " + "-" * (W - 2))

    tot_open = tot_closed = tot_wins = 0
    tot_today = tot_alltime = tot_unreal = 0.0

    for cat in CATEGORIES:
        s = stats.get(cat, {})
        o  = s.get("open", 0)
        cl = s.get("closed_alltime", 0)
        wr = s.get("win_rate")
        rt = s.get("realized_today", 0.0)
        ra = s.get("realized_alltime", 0.0)
        ur = s.get("unrealized", 0.0)

        print(f"  {cat:<18} {o:>5}  {cl:>6}  {_fmt_pct(wr):>6}  "
              f"{_fmt_money(rt):>13}  {_fmt_money(ra):>15}  {_fmt_money(ur):>11}")

        tot_open   += o
        tot_closed += cl
        if wr is not None:
            tot_wins += round(wr * cl)
        tot_today  += rt
        tot_alltime += ra
        tot_unreal  += ur

    overall_wr = (tot_wins / tot_closed) if tot_closed > 0 else None

    print("  " + "-" * (W - 2))
    print(f"  {'TOTAL':<18} {tot_open:>5}  {tot_closed:>6}  {_fmt_pct(overall_wr):>6}  "
          f"{_fmt_money(tot_today):>13}  {_fmt_money(tot_alltime):>15}  "
          f"{_fmt_money(tot_unreal):>11}")
    print("=" * W + "\n")


# ── Excel helpers ──────────────────────────────────────────────────────────────

def _write_header(ws, row: int, cols: list[str]) -> None:
    for c, label in enumerate(cols, 1):
        cell = ws.cell(row=row, column=c, value=label)
        cell.font      = Font(bold=True, color="FFFFFF", size=11)
        cell.fill      = _fill(_CLR_HEADER)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = _border()


def _write_total_row(ws, row: int, values: list, n_cols: int) -> None:
    for c, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=c, value=val)
        cell.font      = Font(bold=True, color="FFFFFF", size=11)
        cell.fill      = _fill(_CLR_TOTAL)
        cell.alignment = Alignment(horizontal="right" if c > 1 else "left",
                                   vertical="center")
        cell.border    = _border()
        if isinstance(val, (int, float)) and c > 1:
            pf = _pnl_fill(val)
            if pf:
                cell.fill = pf
                cell.font = Font(bold=True, size=11,
                                 color=_CLR_GREEN_DARK if val > 0 else _CLR_RED_DARK)


def _apply_pnl_style(cell, value: float) -> None:
    pf = _pnl_fill(value)
    if pf:
        cell.fill = pf
        cell.font = Font(color=_CLR_GREEN_DARK if value > 0 else _CLR_RED_DARK)


def _auto_width(ws) -> None:
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                max_len = max(max_len, len(str(cell.value or "")))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 4, 40)


# ── Sheet builders ─────────────────────────────────────────────────────────────

def _build_dashboard(wb: Workbook, stats: dict[str, dict]) -> None:
    ws = wb.create_sheet("Dashboard")
    ws.sheet_view.showGridLines = False

    headers = ["Category", "Open", "Closed", "Win %",
               "Realized Today", "Realized All-Time", "Unrealized", "Total P&L"]
    _write_header(ws, 1, headers)

    tot_open = tot_closed = tot_wins = 0
    tot_today = tot_alltime = tot_unreal = 0.0

    for r, cat in enumerate(CATEGORIES, 2):
        s  = stats.get(cat, {})
        o  = s.get("open", 0)
        cl = s.get("closed_alltime", 0)
        wr = s.get("win_rate")
        rt = round(s.get("realized_today", 0.0), 2)
        ra = round(s.get("realized_alltime", 0.0), 2)
        ur = round(s.get("unrealized", 0.0), 2)
        total_pnl = round(ra + ur, 2)

        row_fill = _fill(_CLR_ALT_ROW) if r % 2 == 0 else None
        row_data = [cat, o, cl, wr, rt, ra, ur, total_pnl]

        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.border    = _border()
            cell.alignment = Alignment(horizontal="right" if c > 1 else "left",
                                       vertical="center")
            if row_fill and c not in (5, 6, 7, 8):
                cell.fill = row_fill

            if c == 4:   # Win % — None written as empty cell, float as percentage
                if isinstance(val, float):
                    cell.number_format = "0.0%"
                cell.alignment = Alignment(horizontal="center")
            elif c in (5, 6, 7, 8):   # P&L columns
                cell.number_format = '#,##0.00;[RED]-#,##0.00'
                if isinstance(val, (int, float)):
                    _apply_pnl_style(cell, val)

        tot_open    += o
        tot_closed  += cl
        if wr is not None:
            tot_wins += round(wr * cl)
        tot_today   += rt
        tot_alltime += ra
        tot_unreal  += ur

    overall_wr  = (tot_wins / tot_closed) if tot_closed > 0 else None
    tot_total   = round(tot_alltime + tot_unreal, 2)
    total_row_r = len(CATEGORIES) + 2

    total_vals = ["TOTAL", tot_open, tot_closed, overall_wr,
                  round(tot_today, 2), round(tot_alltime, 2),
                  round(tot_unreal, 2), tot_total]
    for c, val in enumerate(total_vals, 1):
        cell = ws.cell(row=total_row_r, column=c, value=val)
        cell.font      = Font(bold=True, size=11)
        cell.border    = _border()
        cell.alignment = Alignment(horizontal="right" if c > 1 else "left",
                                   vertical="center")
        cell.fill      = _fill(_CLR_TOTAL)
        cell.font      = Font(bold=True, color="FFFFFF", size=11)
        if c == 4:
            if isinstance(val, float):
                cell.number_format = "0.0%"
            cell.alignment = Alignment(horizontal="center")
        elif c in (5, 6, 7, 8) and isinstance(val, (int, float)):
            cell.number_format = '#,##0.00;[RED]-#,##0.00'
            pf = _pnl_fill(val)
            if pf:
                cell.fill = pf
                cell.font = Font(bold=True, size=11,
                                 color=_CLR_GREEN_DARK if val > 0 else _CLR_RED_DARK)

    ws.row_dimensions[1].height = 22
    _auto_width(ws)


def _build_open_positions(wb: Workbook, pos_pnl: dict[str, float]) -> None:
    ws = wb.create_sheet("Open Positions")
    ws.sheet_view.showGridLines = False

    headers = ["Category", "Code", "Strategy", "Opened At", "Days Open",
               "Unrealized P&L", "TP Target", "CL Target", "Source"]
    _write_header(ws, 1, headers)

    all_trades = get_all_trades()
    open_trades = sorted(
        [t for t in all_trades if t.get("status") == "open"],
        key=lambda t: (classify_trade(t), str(t.get("opened_at", "")))
    )

    today = date.today()
    r = 2
    for trade in open_trades:
        cat      = classify_trade(trade)
        code     = trade.get("stock_code", "")
        strategy = trade.get("strategy", "")
        opened   = str(trade.get("opened_at", ""))[:19]
        try:
            days_open = (today - date.fromisoformat(str(trade.get("opened_at", ""))[:10])).days
        except Exception:
            days_open = ""

        # Unrealized P&L
        unreal = 0.0
        if code in pos_pnl:
            unreal = pos_pnl[code]
        elif trade.get("trade_type") == "options":
            leg_vals = [pos_pnl.get(l.get("code", "")) for l in trade.get("legs", [])]
            if all(v is not None for v in leg_vals) and leg_vals:
                unreal = sum(leg_vals)

        # TP / CL estimates
        tp_v = cl_v = ""
        if trade.get("trade_type") == "options":
            net_c = float(trade.get("net_credit_per_spread") or 0)
            num_c = float(trade.get("num_contracts") or 1)
            mult  = float(trade.get("multiplier") or 100)
            _CREDIT = {"Bull Put Spread", "Bear Call Spread", "Iron Condor",
                       "Cash-Secured Put", "Covered Call"}
            _DEBIT  = {"Bull Call Spread", "Bear Put Spread"}
            if strategy in _CREDIT and net_c > 0:
                tp_v = round(net_c * (1 - TP_CREDIT_REMAINING) * num_c * mult, 2)
                cl_v = round(-net_c * (CL_CREDIT_MULT - 1) * num_c * mult, 2)
            elif strategy in _DEBIT and net_c < 0:
                from agent.constants import TP_DEBIT_MULT, CL_DEBIT_REMAINING
                debit = abs(net_c)
                tp_v = round(debit * (TP_DEBIT_MULT - 1) * num_c * mult, 2)
                cl_v = round(-debit * (1 - CL_DEBIT_REMAINING) * num_c * mult, 2)
        else:
            entry = float(trade.get("limit_price") or 0)
            qty   = float(trade.get("qty") or 0)
            side  = trade.get("side", "BUY")
            dirn  = 1 if side == "BUY" else -1
            if entry and qty:
                tp_v = round(entry * STOCK_TP_PCT * qty * dirn, 2)
                cl_v = round(-entry * STOCK_CL_PCT * qty * dirn, 2)

        # Source label
        scan_src  = trade.get("scan_source", "")
        wheel_t   = trade.get("wheel_type", "")
        pol_name  = trade.get("_politician", "")
        if scan_src == "politician" and pol_name:
            source = f"Politician: {pol_name}"
        elif scan_src == "russell":
            source = "Russell 2000"
        elif wheel_t:
            source = f"Wheel ({wheel_t})"
        else:
            source = "Main scan"

        row_fill = _fill(_CLR_ALT_ROW) if r % 2 == 0 else None
        row_data = [cat, code, strategy, opened, days_open, unreal, tp_v, cl_v, source]

        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.border    = _border()
            cell.alignment = Alignment(horizontal="right" if c in (5, 6, 7, 8) else "left",
                                       vertical="center")
            if row_fill:
                cell.fill = row_fill
            if c == 6 and isinstance(val, (int, float)):   # Unrealized P&L
                cell.number_format = '#,##0.00;[RED]-#,##0.00'
                _apply_pnl_style(cell, val)
            elif c in (7, 8) and isinstance(val, (int, float)):
                cell.number_format = '#,##0.00;[RED]-#,##0.00'
        r += 1

    if r == 2:
        ws.cell(row=2, column=1, value="No open trades")

    ws.row_dimensions[1].height = 22
    _auto_width(ws)


def _build_trade_history(wb: Workbook) -> None:
    ws = wb.create_sheet("Trade History")
    ws.sheet_view.showGridLines = False

    headers = ["Category", "Code", "Strategy", "Market",
               "Opened At", "Closed At", "Duration (days)", "Realized P&L",
               "Close Reason", "Win/Loss"]
    _write_header(ws, 1, headers)

    all_trades = get_all_trades()
    closed = sorted(
        [t for t in all_trades if t.get("status") == "closed"],
        key=lambda t: str(t.get("closed_at", "")),
        reverse=True,
    )

    for r, trade in enumerate(closed, 2):
        cat      = classify_trade(trade)
        code     = trade.get("stock_code", "")
        strategy = trade.get("strategy", "")
        market   = trade.get("market", "")
        opened   = str(trade.get("opened_at", ""))[:19]
        closed_at = str(trade.get("closed_at", ""))[:19]
        pnl      = float(trade.get("close_pnl") or 0)
        reason   = str(trade.get("close_reason", "")).replace("_", " ")

        try:
            d_open  = date.fromisoformat(str(trade.get("opened_at", ""))[:10])
            d_close = date.fromisoformat(str(trade.get("closed_at", ""))[:10])
            duration = (d_close - d_open).days
        except Exception:
            duration = ""

        win_label = "WIN" if pnl > 0 else ("LOSS" if pnl < 0 else "BREAK")

        row_fill = _fill(_CLR_ALT_ROW) if r % 2 == 0 else None
        row_data = [cat, code, strategy, market, opened, closed_at,
                    duration, pnl, reason, win_label]

        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.border    = _border()
            cell.alignment = Alignment(
                horizontal="right" if c in (7, 8) else "center" if c == 10 else "left",
                vertical="center",
            )
            if row_fill and c != 8 and c != 10:
                cell.fill = row_fill
            if c == 8:   # Realized P&L
                cell.number_format = '#,##0.00;[RED]-#,##0.00'
                _apply_pnl_style(cell, pnl)
            elif c == 10:  # Win/Loss
                wl_color = (_CLR_WIN if pnl > 0
                            else _CLR_LOSS if pnl < 0
                            else _CLR_BREAK)
                cell.fill      = _fill(wl_color)
                cell.font      = Font(bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center")

    if not closed:
        ws.cell(row=2, column=1, value="No closed trades yet")

    ws.row_dimensions[1].height = 22
    _auto_width(ws)


def _build_monthly_trend(wb: Workbook) -> None:
    ws = wb.create_sheet("Monthly Trend")
    ws.sheet_view.showGridLines = False

    headers = ["Month"] + CATEGORIES + ["Total"]
    _write_header(ws, 1, headers)

    by_month = compute_monthly_trend()

    if not by_month:
        ws.cell(row=2, column=1, value="No closed trades yet")
        return

    col_totals = {cat: 0.0 for cat in CATEGORIES}
    grand_total = 0.0

    for r, (month, cat_pnl) in enumerate(by_month.items(), 2):
        row_total = sum(cat_pnl.get(cat, 0.0) for cat in CATEGORIES)
        row_fill  = _fill(_CLR_ALT_ROW) if r % 2 == 0 else None

        ws.cell(row=r, column=1, value=month).border = _border()
        if row_fill:
            ws.cell(row=r, column=1).fill = row_fill

        for c, cat in enumerate(CATEGORIES, 2):
            val  = round(cat_pnl.get(cat, 0.0), 2)
            cell = ws.cell(row=r, column=c, value=val)
            cell.border         = _border()
            cell.number_format  = '#,##0.00;[RED]-#,##0.00'
            cell.alignment      = Alignment(horizontal="right", vertical="center")
            _apply_pnl_style(cell, val)
            col_totals[cat] = round(col_totals[cat] + val, 2)

        # Row total
        rt_cell = ws.cell(row=r, column=len(CATEGORIES) + 2, value=round(row_total, 2))
        rt_cell.border        = _border()
        rt_cell.number_format = '#,##0.00;[RED]-#,##0.00'
        rt_cell.alignment     = Alignment(horizontal="right", vertical="center")
        rt_cell.font          = Font(bold=True)
        _apply_pnl_style(rt_cell, row_total)
        grand_total = round(grand_total + row_total, 2)

    # Totals row
    total_r = len(by_month) + 2
    ws.cell(row=total_r, column=1, value="TOTAL").font = Font(bold=True, color="FFFFFF")
    ws.cell(row=total_r, column=1).fill   = _fill(_CLR_TOTAL)
    ws.cell(row=total_r, column=1).border = _border()

    for c, cat in enumerate(CATEGORIES, 2):
        val  = col_totals[cat]
        cell = ws.cell(row=total_r, column=c, value=val)
        cell.border        = _border()
        cell.number_format = '#,##0.00;[RED]-#,##0.00'
        cell.alignment     = Alignment(horizontal="right", vertical="center")
        cell.font          = Font(bold=True)
        pf = _pnl_fill(val)
        if pf:
            cell.fill = pf
            cell.font = Font(bold=True, color=_CLR_GREEN_DARK if val > 0 else _CLR_RED_DARK)
        else:
            cell.fill = _fill(_CLR_TOTAL)
            cell.font = Font(bold=True, color="FFFFFF")

    gt_cell = ws.cell(row=total_r, column=len(CATEGORIES) + 2, value=grand_total)
    gt_cell.border        = _border()
    gt_cell.number_format = '#,##0.00;[RED]-#,##0.00'
    gt_cell.alignment     = Alignment(horizontal="right", vertical="center")
    pf = _pnl_fill(grand_total)
    if pf:
        gt_cell.fill = pf
        gt_cell.font = Font(bold=True,
                            color=_CLR_GREEN_DARK if grand_total > 0 else _CLR_RED_DARK)
    else:
        gt_cell.fill = _fill(_CLR_TOTAL)
        gt_cell.font = Font(bold=True, color="FFFFFF")

    ws.row_dimensions[1].height = 22
    _auto_width(ws)


# ── Public entry points ────────────────────────────────────────────────────────

def generate_category_report(
    account_us: dict | None,
    account_hk: dict | None,
    monitor_results: list[dict] | None = None,
    report_date: date | None = None,
) -> str | None:
    """
    Generate the category P&L Excel report and persist a daily snapshot.

    When monitor_results is provided (called right after run_monitor), unrealized
    P&L comes from the already-computed eval results — zero extra API calls.
    When called standalone (category_report command), fetches live prices from Alpaca.

    Returns the file path on success, None on failure.
    """
    if report_date is None:
        report_date = date.today()

    # Fetch live pos_pnl only when monitor_results are not available
    pos_pnl: dict[str, float] = {}
    if monitor_results is None:
        pos_pnl = _fetch_pos_pnl(account_us, account_hk)

    stats = compute_category_stats(monitor_results=monitor_results, pos_pnl=pos_pnl,
                                   hkd_rate=HKD_USD_RATE)

    # Console output (same data as Dashboard sheet)
    print_category_summary(stats)

    # Persist daily snapshot
    _append_pnl_history(stats, report_date)

    # Build Excel workbook
    os.makedirs(_REPORTS_DIR, exist_ok=True)
    out_path = os.path.join(_REPORTS_DIR, f"CategoryPnL_{report_date}.xlsx")

    try:
        wb = Workbook()
        # Remove default empty sheet
        wb.remove(wb.active)

        _build_dashboard(wb, stats)
        _build_open_positions(wb, pos_pnl)
        _build_trade_history(wb)
        _build_monthly_trend(wb)

        wb.save(out_path)
        print(f"  Category report saved: {out_path}")
        return out_path
    except Exception as e:
        print(f"  ERROR writing category report: {e}")
        import traceback; traceback.print_exc()
        return None
